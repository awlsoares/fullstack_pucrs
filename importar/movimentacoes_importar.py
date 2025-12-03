# importar/importar.py
from openpyxl import load_workbook
from datetime import datetime
from models import db, Movimentacao

def importar_movimentacoes(caminho_excel):
    wb = load_workbook(caminho_excel, data_only=True)
    sheet = wb['plan_movimentacoes']

    meses = {
        'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
        'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
    }

    print(sheet)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_excel, descricao, forma_pgto, status_pagamento, competencia, tipo, categoria, entrada_valor, saida_valor = row[:9]

        # DATA
        if isinstance(data_excel, datetime):
            data = data_excel.date()
        else:
            data = datetime.strptime(data_excel, "%d/%m/%Y").date()

        # COMPETÊNCIA
        if isinstance(competencia, datetime):
            competencia_str = competencia.strftime('%b/%y').lower()
        else:
            competencia_str = str(competencia).strip().lower()

        mes_abrev, ano_abrev = competencia_str.split('/')
        mes = meses.get(mes_abrev[:3], data.month)
        ano = 2000 + int(ano_abrev) if int(ano_abrev) < 100 else int(ano_abrev)

        # OPERAÇÃO e VALOR
        if entrada_valor and not saida_valor:
            operacao = 'Entrada'
            valor = entrada_valor
        elif saida_valor and not entrada_valor:
            operacao = 'Saída'
            valor = saida_valor
        else:
            continue

        # Converte valor
        if isinstance(valor, str):
            valor = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
        valor = float(valor)

        mov = Movimentacao(
            operacao=operacao,
            data=data,
            descricao=descricao,
            forma_de_pagamento=forma_pgto,
            status_pagamento=status_pagamento,
            mes=mes,
            ano=ano,
            movimentacao_tipo=tipo,
            movimentacao_categoria=categoria,
            valor=valor,
            observacao=''
        )
        print(mov)

        db.session.add(mov)

        db.session.commit()
