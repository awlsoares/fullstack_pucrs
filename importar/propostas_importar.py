# importar/importar.py
from openpyxl import load_workbook
from datetime import datetime, date
from models import db, Cliente, Proposta

def importar_propostas(caminho_excel):
    print("Entrou em importar_propostas!")
    wb = load_workbook(caminho_excel, data_only=True)
    sheet = wb['plan_propostas']
    print(sheet)

    # Começa na segunda linha (pula cabeçalho)
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        print(f"Linha {idx}: {row}")

        numero_proposta = str(row[0])  # Nº DA PROPOSTA
        data_envio_excel = row[1]  # DATA ENVIO
        nome_empresa = row[2]  # EMPRESA
        escopo = row[3]  # ESCOPO
        valor_assinatura = row[4]  # ASSINATURA
        valor_protocolo = row[5]  # PROTOCOLO
        valor_conclusao = row[6]  # CONCLUSÃO
        status = row[7]  # STATUS
        multiplo = row[8] # MÚLTIPLOS CONTRATOS s/n

        if (multiplo == 's'): 
            multiplo = True 
        else:
            multiplo = False

        # Converte data_envio
        if isinstance(data_envio_excel, datetime):
            data_envio = data_envio_excel.date()
        elif isinstance(data_envio_excel, str):
            data_envio = datetime.strptime(data_envio_excel.strip(), "%d/%m/%Y").date()
        else:
            data_envio = date.today()

        # Converte valores monetários (float)
        def to_float(valor):
            if valor is None:
                return 0.0
            if isinstance(valor, (int, float)):
                return float(valor)
            valor = str(valor)
            valor = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
            return float(valor) if valor else 0.0

        valor_assinatura = to_float(valor_assinatura)
        valor_protocolo = to_float(valor_protocolo)
        valor_conclusao = to_float(valor_conclusao)

        # print(f"Assinatura {valor_assinatura}, Protocolo {valor_protocolo}, Aprovação {valor_aprovacao}.")

        # Busca o cliente pelo nome
        cliente = Cliente.query.filter(Cliente.nome.ilike(nome_empresa.strip())).first()
        if not cliente:
            print(f"Cliente '{nome_empresa}' não encontrado. Pulando proposta {numero_proposta}.")
            continue

        # print("Cliente: ", cliente, " ID: ", cliente.id, type(cliente.id))
        # print("numero_proposta: ", numero_proposta, type(numero_proposta))
        # print("Data do envio: ", data_envio, type(data_envio))
        # print("Valor Assinatura: ", valor_assinatura, type(valor_assinatura))
        # print("Escopo: ", escopo, type(escopo))
        # print("Status: ", status, type(status))


        # Cria objeto Proposta
        try:
            proposta = Proposta(
                numero_proposta = str(numero_proposta),
                data_envio = data_envio,
                cliente_id = cliente.id,
                escopo = escopo or '',
                valor_assinatura  =valor_assinatura,
                valor_protocolo = valor_protocolo,
                valor_conclusao = valor_conclusao,
                status_proposta = status or 'Aguardando Aprovação',
                observacao = '',  # pode preencher se tiver coluna extra
                versao = 0,
                contratos_multiplos = multiplo
            )
            db.session.add(proposta)
            print(f"Linha {idx}: Proposta criada com sucesso.")
        except Exception as e:
            db.session.rollback()
            print(f"Erro ao criar proposta na linha {idx}: {e}")

    # Commit geral no final
    try:
        db.session.commit()
        print("Todas as propostas foram salvas com sucesso.")
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao salvar propostas: {e}")