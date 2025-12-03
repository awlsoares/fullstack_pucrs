from openpyxl import load_workbook
from models import db, Cliente

def importar_clientes(caminho_excel):
    print("Entrou em importar_clientes!")
    wb = load_workbook(caminho_excel, data_only=True)
    sheet = wb['plan_clientes']
    print(sheet)

    # Começa na segunda linha (pula cabeçalho)
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        print(f"Linha {idx}: {row}")

        nome_empresa = row[0]  # EMPRESA
        cnpj = row[1]
        observacao = row[2]

        # Cria objeto Cliente
        try:
            cliente = Cliente(
                nome = nome_empresa,
                cnpj = cnpj,
                observacao = observacao
            )
            db.session.add(cliente)
            print(f"Linha {idx}: Cliente criado com sucesso.")
        except Exception as e:
            db.session.rollback()
            print(f"Erro ao criar cliente na linha {idx}: {e}")

    # Commit geral no final
    try:
        db.session.commit()
        print("Todas os clientes foram salvos com sucesso.")
    except Exception as e:
        db.session.rollback()
        print(f"Erro ao salvar clientes: {e}")